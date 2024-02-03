from time import sleep
from datetime import datetime

import openpyxl
import pandas as pd
from alive_progress import alive_bar

database_lock = False


def get_column_names(cursor, table_name):
    cursor.execute(f"PRAGMA table_info({table_name})")
    return [col_info[1] for col_info in cursor.fetchall()]


def write_to_db(c, conn, match_list, overwrite=True):
    if not overwrite:
        # Filter out matches that already exist in the database
        new_match_list = []
        skipped_matches = 0
        with alive_bar(
            len(match_list), title="Checking for matches that already exist"
        ) as bar:
            for match in match_list:
                # Parameterized query to check if a match exists in the database
                query = (
                    "SELECT COUNT(*) FROM tennis_matches WHERE "
                    "match_id LIKE ? AND "
                    "((A_name = ? AND B_name = ?) OR "
                    "(A_name = ? AND B_name = ?)) AND "
                    "round = ?"
                )
                match_id_pattern = (
                    f"{match['match_id'][:4]}%{match['match_id'].split('_')[1]}%"
                )
                params = (
                    match_id_pattern,
                    match["A_name"],
                    match["B_name"],
                    match["B_name"],
                    match["A_name"],
                    match["round"],
                )
                c.execute(query, params)
                exists = c.fetchone()[0] > 0

                max_match_id = match["match_id"] + "_" + "000"

                # get max of match_id that starts with match['match_id']
                c.execute(
                    f"SELECT MAX(match_id) FROM tennis_matches WHERE match_id LIKE '{match['match_id']}%'"
                )
                fetched_max_match_id = c.fetchone()[0]
                if fetched_max_match_id:
                    max_match_id = fetched_max_match_id

                # Also check new_match_list for matches that already exist
                for m in new_match_list:
                    if (
                        # if match_id starts with match['match_id']
                        m["match_id"].startswith(match["match_id"])
                        and m["match_id"] > max_match_id
                    ):
                        max_match_id = m["match_id"]

                # if exists, add 1 to match_id
                if max_match_id:
                    match["match_id"] += "_" + str(
                        int(max_match_id.split("_")[-1]) + 1
                    ).zfill(3)

                if not exists:
                    new_match_list.append(match)
                else:
                    skipped_matches += 1

                bar()

        match_list = new_match_list
        print(f"Skipped {skipped_matches} matches that already exist in the database.")

    # if overwrite delete all matches from the tournament
    tourney_ids = list(
        set(["_".join(m["match_id"].split("_")[:2]) for m in match_list])
    )

    if overwrite:
        print("Deleting existing data from the database.")
        if tourney_ids:  # Make sure there is at least one id to avoid a malformed query
            query = "DELETE FROM tennis_matches WHERE " + " OR ".join(
                f"match_id LIKE '{tourney_id[:4]}%{tourney_id.split('_')[1]}%'"
                for tourney_id in tourney_ids
            )
            c.execute(query)
            conn.commit()

    new_data = pd.DataFrame(match_list)
    existing_columns = get_column_names(c, "tennis_matches")

    print("Writing new data to the database.")
    for key in new_data.columns:
        if key not in existing_columns:
            c.execute(f"ALTER TABLE tennis_matches ADD COLUMN {key} TEXT")
            conn.commit()

    new_data.to_sql("tennis_matches", conn, if_exists="append", index=False)

    print(f"Database updated with data from {len(match_list)} matches.")


# temp function to delete all matches with match_id like '20240123'
def delete_tourney(c, conn, year, tourney_id):
    query = f"DELETE FROM tennis_matches WHERE match_id LIKE '{year}%{tourney_id}%'"
    c.execute(query)
    conn.commit()


def write_to_pd(fixture_list):
    prediction_tourney_data = []

    if not fixture_list:
        return

    print("Writing new data to the Excel file.")

    # Load existing data from the Excel file
    existing_data = pd.read_excel("data/predictions.xlsx", sheet_name="Predictions")

    fixtures_added_to_excel = 0

    for m in fixture_list:
        # Check if the match is already in the existing data
        tourney_id = "_".join(m["match_id"].split("_")[:2])
        is_existing = (
            (
                existing_data["match_id"].apply(lambda x: "_".join(x.split("_")[:2]))
                == tourney_id
            )
            & (
                (
                    (existing_data["A_name"] == m["A_name"])
                    & (existing_data["B_name"] == m["B_name"])
                )
                | (
                    (existing_data["A_name"] == m["B_name"])
                    & (existing_data["B_name"] == m["A_name"])
                )
            )
            & (existing_data["round"] == m["round"])
        )

        if not is_existing.any():
            # The match does not exist in the Excel file, so add it to the list
            prediction_match = {}
            prediction_match["tourney_date"] = m["match_id"].split("_")[0]
            prediction_match["match_id"] = m["match_id"]
            prediction_match["A_name"] = m["A_name"]
            prediction_match["B_name"] = m["B_name"]
            prediction_match["surface"] = m["tourney_surface"]
            prediction_match["tourney_level"] = m["tourney_level"]
            prediction_match["round"] = m["round"]
            prediction_match["tourney_name"] = m["tourney_name"]
            prediction_match["tourney_IOC"] = m["tourney_IOC"]

            prediction_tourney_data.append(prediction_match)
            fixtures_added_to_excel += 1

    prediction_tourney_data = pd.DataFrame(prediction_tourney_data)

    book = openpyxl.load_workbook("data/predictions.xlsx")
    ws = book["Predictions"]
    last_row = ws.max_row

    for i, row in prediction_tourney_data.iterrows():
        for j, item in enumerate(row):
            ws.cell(row=i + last_row + 1, column=j + 1, value=item)

    book.save("data/predictions.xlsx")

    print(f"Excel file updated with data from {fixtures_added_to_excel} matches.")
