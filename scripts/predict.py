import os
import sqlite3
import sys

import jellyfish
import joblib
import numpy as np
import openpyxl
import pandas as pd
from alive_progress import alive_bar
from keras.models import load_model

# Add the parent directory of 'scripts' to sys.path
parent_dir = os.path.abspath(
    os.path.join(os.path.dirname(os.path.dirname(__file__)), ".")
)
sys.path.append(parent_dir)

from config.config import STAT_LABELS
from scripts.data_helpers import simplify_name
from scripts.stats import get_all_matches_for_player, get_match_stats

labels = STAT_LABELS


def find_latest_model():
    model_dir = "models/NN_model"
    models = [f for f in os.listdir(model_dir) if f.endswith(".sav")]
    models.sort(
        key=lambda x: os.path.getmtime(os.path.join(model_dir, x)), reverse=True
    )
    print(os.path.join(model_dir, models[0]))
    return os.path.join(model_dir, models[0])


def find_latest_transformer():
    transformer_dir = "models/transformer"
    transformers = [f for f in os.listdir(transformer_dir) if f.endswith(".sav")]
    transformers.sort(
        key=lambda x: os.path.getmtime(os.path.join(transformer_dir, x)), reverse=True
    )
    print(os.path.join(transformer_dir, transformers[0]))
    return os.path.join(transformer_dir, transformers[0])


def prepare_data(df, conn):
    player_match_dict = {}
    X = {"AB": [], "BA": []}
    stats_list = []  # Added to store all calculated stats

    # Count matches where "p" is blank or null
    num_matches_to_predict = len(
        df[(df["p"] == "") | (df["p"].isnull())]
    )  # Matches to predict

    with alive_bar(num_matches_to_predict) as bar:
        for index, row in df.iterrows():
            # Skip rows that already have a p value
            if row["p"] != "" and pd.notnull(row["p"]):
                continue

            player_A = simplify_name(row["A_name"])
            player_B = simplify_name(row["B_name"])

            if player_A not in player_match_dict:
                player_A_matches = get_all_matches_for_player(player_A, conn)
                if (
                    not player_A_matches
                ):  # If no matches found, find the best match name in the database.
                    player_A = find_best_name_match(player_A, conn)
                    player_A_matches = get_all_matches_for_player(player_A, conn)
                player_match_dict[player_A] = player_A_matches
            else:
                player_A_matches = player_match_dict[player_A]

            if player_B not in player_match_dict:
                player_B_matches = get_all_matches_for_player(player_B, conn)
                if (
                    not player_B_matches
                ):  # If no matches found, find the best match name in the database.
                    player_B = find_best_name_match(player_B, conn)
                    player_B_matches = get_all_matches_for_player(player_B, conn)
                player_match_dict[player_B] = player_B_matches
            else:
                player_B_matches = player_match_dict[player_B]

            player_A_matches = [
                player_A_matches[i]
                for i in range(len(player_A_matches))
                if player_A_matches[i]["match_id"] < row["match_id"]
            ]

            player_B_matches = [
                player_B_matches[i]
                for i in range(len(player_B_matches))
                if player_B_matches[i]["match_id"] < row["match_id"]
            ]

            data_list = get_match_stats(
                row, player_A, player_B, player_A_matches, player_B_matches
            )

            X["AB"].append(data_list)
            stats_list.append({"index": index, "stats": data_list})

            # Swap player A and player B for BA data
            row_swapped = swap_players(row)
            data_list_swapped = get_match_stats(
                row_swapped, player_B, player_A, player_B_matches, player_A_matches
            )

            X["BA"].append(data_list_swapped)

            bar()

    # Convert lists to np.array
    X["AB"] = np.array(X["AB"])
    X["BA"] = np.array(X["BA"])

    return X, stats_list  # Return stats_list


def predict(model, X):
    return model.predict(X, verbose=0)


def find_best_name_match(name, conn):
    cursor = conn.cursor()
    # Replace 'your_table_name' with the appropriate table name.
    cursor.execute("SELECT A_simplified_name, B_simplified_name FROM tennis_matches")
    names_in_db = set(
        [row[0] for row in cursor.fetchall()] + [row[1] for row in cursor.fetchall()]
    )

    best_match = None
    best_score = 0

    for db_name in names_in_db:
        score = jellyfish.jaro_winkler(name, db_name)
        if score > best_score:
            best_score = score
            best_match = db_name

    return best_match


def load_data():
    # If excel file is already open, return empty dataframe
    try:
        data_path = "data/predictions.xlsx"
        book = openpyxl.load_workbook(data_path, read_only=True)
        ws = book["Predictions"]
        data = ws.values
        headers = next(data)[0:]
        df = pd.DataFrame(data, columns=headers)
        book.close()
    except PermissionError:
        print("Excel file is already open.")
        df = pd.DataFrame()
    return df


def swap_players(row):
    # Swap player A and player B
    row_swapped = row.copy()
    player_A_columns = [col for col in row.index if col.startswith("A_")]
    player_B_columns = [col for col in row.index if col.startswith("B_")]

    for col_A, col_B in zip(player_A_columns, player_B_columns):
        row_swapped[col_A], row_swapped[col_B] = row[col_B], row[col_A]

    return row_swapped


def save_predictions(df, predictions, stats_list):
    data_path = "data/predictions.xlsx"

    # Load the workbook
    book = openpyxl.load_workbook(data_path)
    ws = book["Predictions"]

    # Update p values
    for index, prediction in zip(df.index[df["p"].isnull()], predictions):
        ws.cell(row=index + 2, column=df.columns.get_loc("p") + 1, value=prediction[0])

    # Save all the stats
    for i, label in enumerate(labels):
        if label not in df.columns:
            ws.cell(row=1, column=len(df.columns) + i + 1, value=label)
            column_index = len(df.columns) + i + 1
        else:
            column_index = df.columns.get_loc(label) + 1

        # Update only the rows with non-null p values
        for stat in stats_list:
            ws.cell(row=stat["index"] + 2, column=column_index, value=stat["stats"][i])

    # Save and close the workbook
    book.save(data_path)
    book.close()


def predict_main():
    latest_model_path = find_latest_model()
    model = load_model(latest_model_path)
    latest_transformer_path = find_latest_transformer()
    NN_transformer = joblib.load(latest_transformer_path)

    print("Models loaded")

    conn = sqlite3.connect("data/matches.sqlite")

    df = load_data()

    X, stats_list = prepare_data(df, conn)  # Receive stats_list

    print("Data loaded")

    # Check if X is empty
    if not X["AB"].any():
        print("No new matches to predict")
        return []

    # Transform both sets of data
    X_transformed = {
        "AB": NN_transformer.transform(X["AB"]),
        "BA": NN_transformer.transform(X["BA"]),
    }

    # Predict for both sets of data
    predictions = {
        "AB": predict(model, X_transformed["AB"]),
        "BA": predict(model, X_transformed["BA"]),
    }

    # Calculate the final prediction as the average of A vs B and 1 - B vs A
    final_predictions = (predictions["AB"] + (1 - predictions["BA"])) / 2

    # Prepare list of matches with predictions
    matches_with_predictions = []
    for index, row in df.iterrows():
        if row["p"] == "" or pd.isnull(row["p"]):
            matches_with_predictions.append(
                {
                    "match_id": row["match_id"],
                    "A_name": row["A_name"],
                    "B_name": row["B_name"],
                    "tourney_name": row["tourney_name"],
                    "round": row["round"],
                    "prediction": final_predictions[len(matches_with_predictions)][0],
                }
            )

    # Save the final predictions
    save_predictions(df, final_predictions, stats_list)

    return matches_with_predictions


if __name__ == "__main__":
    predict_main()
